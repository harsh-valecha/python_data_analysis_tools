import pandas as pd
import openpyxl

file = pd.read_csv('practice.csv')
df = pd.DataFrame(data=file)
# print(df.columns)


# accessing months in the dataframe 
# print(df['Month'][1])

# for i in zip(df['Month'],df['Temperature']):
#     print(i)


# loc to access values 
# print(df.loc[1].iat[1]) # getting 1 row 1st column value

# storing the month and temperatures into list of tuples 
values = list(zip(df.index,df['Month'],df['Temperature']))
# print(values[:5]) # first five values


# writing the data to excel file
wb = openpyxl.load_workbook('temp.xlsx')
wb.create_sheet('temp')
print(wb.sheetnames)

sheet = wb.get_sheet_by_name('temp')

sheet.cell(row=1,column=1).value = 'S.No'
sheet.cell(row=1,column=2).value = 'Month'
sheet.cell(row=1,column=3).value='Temperature'

for i in range(len(values)):
    print(values[i][0],values[i][1],values[i][2])
    # adding serial number
    sheet.cell(row=i+2,column=1).value = values[i][0]
    # adding month
    sheet.cell(row=i+2,column=2).value= values[i][1]
    # adding temperature
    sheet.cell(row=i+2,column=3).value= values[i][2]

wb.save('temp.xlsx')





