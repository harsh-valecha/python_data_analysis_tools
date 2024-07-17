import openpyxl

# wb = openpyxl.Workbook()

# sheet = wb.active

# c1 = sheet.cell(row=1,column=1)
# c1.value = 'first_name'

# c2 = sheet.cell(row=1,column=2)
# c2.value= 'last_name'

# students ={
#     'first_name':['harsh','kamlesh','jaktap'],
#     'last_name':['valecha','chaturvedy','pandey']
# }

# for i in enumerate(students['first_name']):
#     sheet.cell(row=i[0]+2,column=1).value = i[1]

# for i in enumerate(students['last_name']):
#     sheet.cell(row=i[0]+2,column=2).value=i[1]


# wb.save('temp.xlsx')

# reading from excel 
wb = openpyxl.load_workbook('temp.xlsx')
sheet = wb.active
print(sheet.max_column,sheet.max_row)
# printing values 
for i in range(sheet.max_row):
    for j in range(sheet.max_column):
        print(sheet.cell(row=i+1,column=j+1).value,end=' ')
    print()

